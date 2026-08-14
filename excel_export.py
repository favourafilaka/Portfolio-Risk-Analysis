from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule

def write_overview_sheet(workbook, portfolio_data):
    '''
    Writes the Portfolio Overview sheet.
    Parameters:
    workbook (openpyxl.Workbook): Workbook to modify
    portfolio_data (list[dict]): Portfolio data, one dict per ticker with
        keys ticker, weight, portfolio_value, return
    Returns:
    None: Modifies workbook in place
    '''
    sheet = workbook.create_sheet("Portfolio Overview")
    sheet["A1"] = "Portfolio Overview"
    sheet["A3"] = "Ticker"
    sheet["B3"] = "Weight"
    sheet["C3"] = "Portfolio Value"
    sheet["D3"] = "Return"
    for row, data in enumerate(portfolio_data, start=4):
        sheet.cell(row=row, column=1, value=data["ticker"])
        sheet.cell(row=row, column=2, value=data["weight"])
        sheet.cell(row=row, column=3, value=data["portfolio_value"])
        sheet.cell(row=row, column=4, value=data["return"])


def write_risk_sheet(workbook, risk_metrics):
    '''
    Writes the Risk Metrics sheet.
    Parameters:
    workbook (openpyxl.Workbook): Workbook to modify
    risk_metrics (dict): Portfolio risk metrics
    Returns:
    None: Modifies workbook in place
    '''
    sheet = workbook.create_sheet("Risk Metrics")
    sheet["A1"] = "Risk Metrics"
    sheet["A3"] = "Metric"
    sheet["B3"] = "Value"
    metrics = [
        ("Volatility", risk_metrics["volatility"]),
        ("Sharpe Ratio", risk_metrics["sharpe"]),
        ("Sortino Ratio", risk_metrics["sortino"]),
        ("VaR", risk_metrics["var"]),
        ("CVaR", risk_metrics["cvar"]),
        ("Maximum Drawdown", risk_metrics["drawdown"]),
    ]
    for row, (metric, value) in enumerate(metrics, start=4):
        sheet.cell(row=row, column=1, value=metric)
        sheet.cell(row=row, column=2, value=value)
    sheet.conditional_formatting.add(
        "B4:B9",
        CellIsRule(operator="lessThan", formula=["0"])
    )


def write_benchmark_sheet(workbook, benchmark_data):
    '''
    Writes the Benchmark Comparison sheet.
    Parameters:
    workbook (openpyxl.Workbook): Workbook to modify
    benchmark_data (dict): Benchmark comparison data
    Returns:
    None: Modifies workbook in place
    '''
    sheet = workbook.create_sheet("Benchmark Comparison")
    sheet["A1"] = "Benchmark Comparison"
    sheet["A3"] = "Metric"
    sheet["B3"] = "Value"
    metrics = [
        ("Beta", benchmark_data["beta"]),
        ("Alpha", benchmark_data["alpha"]),
        ("Tracking Error", benchmark_data["tracking_error"]),
        ("Information Ratio", benchmark_data["information_ratio"]),
    ]
    for row, (metric, value) in enumerate(metrics, start=4):
        sheet.cell(row=row, column=1, value=metric)
        sheet.cell(row=row, column=2, value=value)


def write_backtesting_sheet(workbook, backtest_data):
    '''
    Writes the Backtesting sheet.
    Parameters:
    workbook (openpyxl.Workbook): Workbook to modify
    backtest_data (dict): Backtesting results
    Returns:
    None: Modifies workbook in place
    '''
    sheet = workbook.create_sheet("Backtesting")
    sheet["A1"] = "Backtesting Results"
    sheet["A3"] = "Metric"
    sheet["B3"] = "Value"
    metrics = [
        ("Annualised Return", backtest_data["annualised_return"]),
        ("Calmar Ratio", backtest_data["calmar_ratio"]),
        ("Win Rate", backtest_data["win_rate"]),
    ]
    for row, (metric, value) in enumerate(metrics, start=4):
        sheet.cell(row=row, column=1, value=metric)
        sheet.cell(row=row, column=2, value=value)


def write_monte_carlo_sheet(workbook, simulation_results):
    '''
    Writes the Monte Carlo sheet.
    Parameters:
    workbook (openpyxl.Workbook): Workbook to modify
    simulation_results (dict): Monte Carlo simulation summary — keys mean,
        median, percentile_5, percentile_95 (from
        monte_carlo.calculate_simulation_summary)
    Returns:
    None: Modifies workbook in place
    '''
    sheet = workbook.create_sheet("Monte Carlo")
    sheet["A1"] = "Monte Carlo Simulation"
    sheet["A3"] = "Statistic"
    sheet["B3"] = "Value"
    summary_stats = [
        ("Mean", simulation_results["mean"]),
        ("Median", simulation_results["median"]),
        ("5th Percentile", simulation_results["percentile_5"]),
        ("95th Percentile", simulation_results["percentile_95"]),
    ]
    for row, (statistic, value) in enumerate(summary_stats, start=4):
        sheet.cell(row=row, column=1, value=statistic)
        sheet.cell(row=row, column=2, value=value)


def write_stress_test_sheet(workbook, stress_results):
    '''
    Writes the Stress Test sheet.
    Parameters:
    workbook (openpyxl.Workbook): Workbook to modify
    stress_results (pandas.DataFrame): Stress scenario results from
        stress_testing.generate_stress_summary(), indexed by scenario
        name with columns shocked_value, absolute_impact,
        percentage_impact
    Returns:
    None: Modifies workbook in place
    '''
    sheet = workbook.create_sheet("Stress Testing")
    sheet["A1"] = "Stress Test Results"
    sheet["A3"] = "Scenario"
    sheet["B3"] = "Shocked Value"
    sheet["C3"] = "Absolute Impact"
    sheet["D3"] = "Percentage Impact"

    for row, (scenario_name, data) in enumerate(stress_results.iterrows(), start=4):
        sheet.cell(row=row, column=1, value=scenario_name)
        sheet.cell(row=row, column=2, value=data["shocked_value"])
        sheet.cell(row=row, column=3, value=data["absolute_impact"])
        sheet.cell(row=row, column=4, value=data["percentage_impact"])

    last_row = 3 + len(stress_results)
    sheet.conditional_formatting.add(
        f"D4:D{last_row}",
        CellIsRule(operator="lessThan", formula=["-0.10"])
    )

def write_dashboard_sheet(workbook):
    '''
    Writes the Portfolio Dashboard sheet.
    Parameters:
    workbook (openpyxl.Workbook): Workbook to modify
    Returns:
    None: Modifies workbook in place
    '''
    sheet = workbook.create_sheet("Dashboard")

    sheet["A1"] = "Portfolio Risk Dashboard"

    sheet["A3"] = "Total Portfolio Value"
    sheet["C3"] = "Volatility"
    sheet["E3"] = "Sharpe Ratio"

    sheet["A4"] = "=SUM('Portfolio Overview'!C4:C100)"
    sheet["C4"] = "='Risk Metrics'!B4"
    sheet["E4"] = "='Risk Metrics'!B5"

    sheet["A6"] = "Maximum Drawdown"
    sheet["C6"] = "VaR"
    sheet["E6"] = "Beta"

    sheet["A7"] = "='Risk Metrics'!B9"
    sheet["C7"] = "='Risk Metrics'!B7"
    sheet["E7"] = "='Benchmark Comparison'!B4"

def apply_standard_formatting(worksheet):
    '''
    Applies standard worksheet formatting.
    Parameters:
    worksheet (openpyxl.worksheet.Worksheet): Worksheet to format
    Returns:
    None: Modifies worksheet in place
    '''
    for cell in worksheet[3]:
        cell.font = cell.font.copy(bold=True)
    for column in worksheet.columns:
        column_letter = column[0].column_letter
        worksheet.column_dimensions[column_letter].width = max(
            len(str(cell.value or "")) for cell in column
        ) + 2


def export_to_excel(all_data, output_path):
    '''
    Exports portfolio analysis results to Excel.
    Parameters:
    all_data (dict): Portfolio analysis results — keys portfolio, risk,
        benchmark, backtesting, monte_carlo, stress_testing
    output_path (str): Excel output file path
    Returns:
    None: Saves workbook to specified path
    '''
    workbook = Workbook()
    write_dashboard_sheet(workbook)
    write_overview_sheet(workbook, all_data["portfolio"])
    write_risk_sheet(workbook, all_data["risk"])
    write_benchmark_sheet(workbook, all_data["benchmark"])
    write_backtesting_sheet(workbook, all_data["backtesting"])
    write_monte_carlo_sheet(workbook, all_data["monte_carlo"])
    write_stress_test_sheet(workbook, all_data["stress_testing"])

    for sheet in workbook.worksheets:
        apply_standard_formatting(sheet)
    workbook.save(output_path)
